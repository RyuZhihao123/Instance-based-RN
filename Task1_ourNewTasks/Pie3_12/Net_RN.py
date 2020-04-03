import numpy as np
import tensorflow as tf
from keras.optimizers import SGD,Adam
import matplotlib.pyplot as plt
from  openpyxl import  Workbook
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar, ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse, os, cv2, keras, sklearn
from utils.model_rn import ModelRN



""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000

m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--savedir", default= 'RN')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")
dir_results   = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # backup path. 'results/RN'




def GetInformation(flag):
    if flag == 'train':
        return train_num, config.dir_Charts_train, config.path_groundTruth_train
    if flag == 'val':
        return val_num, config.dir_Charts_val, config.path_groundTruth_val
    if flag == 'test':
        return test_num, config.dir_Charts_test, config.path_groundTruth_test

def LoadChartDataset(flag = 'train'):

    __max_load_num, __dirPath, __path_groundTruth = GetInformation(flag)

    images = []
    for i in range(__max_load_num):
        fileName = __dirPath+config.chartName.format(i)
        images.append(cv2.imread(fileName))

    labels = []
    lines = ReadLinesInFile(__path_groundTruth)

    for line in lines:
        elements = line.split()
        if len(elements) == 0:
            continue
        if len(labels) >= __max_load_num:
            break

        labels.append([float(p) for p in elements])

    print("[Load {}] (images, labels):".format(flag), len(images),len(labels))

    images = np.array(images, dtype='float32')
    labels = np.array(labels, dtype='float32')
    images = (images / 255.)

    print("---------------------------")
    print("[Process] x: ", images.shape)
    print("[Process] y: ", labels.shape, config.max_obj_num)

    return images, labels

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(sess, x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.GetPredictions(sess,x,y,x.shape[0])

    predictFile = open(dir_results+flag+"_predicted_results.txt",'w')
    for i in range(predict_Y.shape[0]):
        for t in range(dim):  # save the ground truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y[:predict_Y.shape[0]] * 100) + .125)

    return MLAE

if __name__ == '__main__':
    ClearDir(dir_results)
    ClearDir(dir_results+"backup")

    x_train, y_train = LoadChartDataset(flag='train')
    x_val, y_val = LoadChartDataset(flag='val')
    x_test, y_test = LoadChartDataset(flag='test')

    x_train -= .5
    x_val -= .5
    x_test -= .5

    print("----------Build network----------------")
    model = ModelRN(learning_rate=a.lr, batch_size=m_batchSize, c_dim=3, a_dim=config.max_obj_num)

    # TensorFlow environment.
    init = tf.global_variables_initializer()
    sess = tf.Session()
    sess.run(init)

    saver = tf.train.Saver(max_to_keep=1)


    print("----------Training-------------------")
    history_batch = []
    history_iter = []
    batch_amount = train_num//m_batchSize
    rest_size = train_num - (batch_amount*m_batchSize)

    # information of the best model on validation set.
    best_val_loss = 99999.99999
    best_model_name = "xxxxx"
    best_train_loss = 99999.99999

    for iter in range(m_epoch):

        # shuffle the training set
        index = [i for i in range(train_num)]
        np.random.shuffle(index)

        for bid in range(batch_amount):

            # using the shuffle index...
            x_batch = x_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]
            y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

            loss = model.Run_one_batch(sess, x_batch, y_batch)

            if bid % m_print_loss_step == 0:
                print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                              GetProcessBar(bid, batch_amount), loss))
                history_batch.append([iter, bid, loss])

        # one epoch is done. Do some information collections.
        train_iter_loss = model.GetTotalLoss(sess,x_train, y_train, x_train.shape[0])
        val_iter_loss =  model.GetTotalLoss(sess,x_val, y_val, x_val.shape[0])
        print("#######>>>> iter({}/{}) train_loss={}, val_loss = {}".format(iter, m_epoch, train_iter_loss,
                                                                             val_iter_loss))
        history_iter.append([iter, train_iter_loss, val_iter_loss])

        # if a.backup == True:   # whether to save the weight after each epoch
        #     saver.save(sess, dir_results + "/backup/model_{}_{}.ckpt".format(iter,val_iter_loss))

        if val_iter_loss < best_val_loss:  # save the best model on Validation set.
            best_model_name = dir_results + "model_RN_onVal_{}.ckpt".format(val_iter_loss)
            saver.save(sess, best_model_name)
            best_val_loss = val_iter_loss
            best_train_loss = train_iter_loss


    print("----------Training Done-------------------")
    saver.restore(sess, best_model_name)

    # test on the testing set.
    test_loss = model.GetTotalLoss(sess, x_test, y_test, x_test.shape[0])

    # Save the predicted results and ground truth.
    MLAE_train = SavePredictedResult(sess, x_train, y_train, 'train')
    MLAE_val = SavePredictedResult(sess, x_val, y_val, 'val')
    MLAE_test = SavePredictedResult(sess, x_test, y_test, 'test')

    # save the training information.
    wb = Workbook()
    ws1 = wb.active                         # MSE/MLAE
    ws1.title = "MLAE_MSE"
    ws2 = wb.create_sheet("EPOCH loss")      # iteration loss
    ws3 = wb.create_sheet("BATCH loss")     # batch loss

    ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
    ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])
    for i in range(len(history_iter)):
        ws2.append(history_iter[i])
    for i in range(len(history_batch)):
        ws3.append(history_batch[i])
    ws1.append(["Train loss", best_train_loss])
    ws1.append(["Val loss", best_val_loss])
    ws1.append(["Test loss", test_loss])
    ws1.append(["Train MLAE", MLAE_train])
    ws1.append(["val MLAE", MLAE_val])
    ws1.append(["Test MLAE", MLAE_test])

    wb.save(dir_results + "train_info.xlsx")

    print("Training MSE:", best_train_loss)
    print("Validat. MSE", best_val_loss)
    print("Testing MSE:", test_loss)
    print("Training MLAE:", MLAE_train)
    print("Validat. MLAE", MLAE_val)
    print("Testing MLAE:", MLAE_test)
