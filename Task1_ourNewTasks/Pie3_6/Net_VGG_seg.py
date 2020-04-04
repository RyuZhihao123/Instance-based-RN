import numpy as np
import keras

from keras.optimizers import SGD,Adam
import matplotlib.pyplot as plt
import cv2
import os
from keras import models
from keras import layers
from openpyxl import Workbook
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar, ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse, sklearn



""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000

m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--savedir", default= 'VGG_seg')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")
dir_results   = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # backup path. 'results/VGG_seg'



def GetInformation(flag):
    if flag == 'train':
        return train_num, config.dir_subCharts_train, config.path_groundTruth_train
    if flag == 'val':
        return val_num, config.dir_subCharts_val, config.path_groundTruth_val
    if flag == 'test':
        return test_num, config.dir_subCharts_test, config.path_groundTruth_test

def LoadSeparateChartDataSet(flag = 'train'):

    __max_load_num, __dirSubPath, __path_groundTruth = GetInformation(flag)

    images = np.ones((__max_load_num, Config.image_height, config.image_width, config.max_obj_num ), dtype='float32')
    count = 0
    os.chdir(__dirSubPath)
    for imgID in range(__max_load_num):
        for barID in range(config.max_obj_num):
            imagePath = config.subChartName.format(imgID, barID)

            if os.path.exists(imagePath):  # if file exists.
                images[imgID,:,:,barID] = cv2.cvtColor(cv2.imread(imagePath), cv2.COLOR_RGB2GRAY) / 255.

        if count % 5000 == 0:
            print("Loaded: {}/{}".format(count, __max_load_num))
        count += 1


    # read labels from file.
    labels = []
    lines = ReadLinesInFile(__path_groundTruth)  #

    for line in lines:
        elements = line.split()
        if len(elements) == 0:
            continue
        if len(labels) >= __max_load_num:
            break
        labels.append([float(p) for p in elements])

    labels = np.array(labels)

    print("---------------------------")
    print("[Process] x: ", images.shape)
    print("[Process] y: ", labels.shape, config.max_obj_num)

    return images, labels




# VGG that is same as Daniel's code.
def VGG():
    feature_generator = keras.applications.VGG19(weights=None, include_top=False, input_shape=(100, 100, config.max_obj_num))

    MLP = models.Sequential()
    MLP.add(layers.Flatten(input_shape=feature_generator.output_shape[1:]))
    MLP.add(layers.Dense(256, activation='relu', input_dim=(100, 100, config.max_obj_num)))
    MLP.add(layers.Dropout(0.5))
    MLP.add(layers.Dense(config.max_obj_num, activation='linear'))  # REGRESSION

    model = keras.Model(inputs=feature_generator.input, outputs=MLP(feature_generator.output))
    return model





# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.predict(x, batch_size=1)
    predictFile = open(dir_results+flag+"_predicted_results.txt",'w')
    for i in range(x.shape[0]):
        for t in range(dim):  # save the ground truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':
    ClearDir(dir_results)
    ClearDir(dir_results+"backup")

    x_train, y_train = LoadSeparateChartDataSet(flag='train')
    x_val, y_val = LoadSeparateChartDataSet(flag='val')
    x_test, y_test = LoadSeparateChartDataSet(flag='test')

    x_train -= .5
    x_val -= .5
    x_test -= .5

    print("----------Build Network----------------")
    model = VGG()
    model.compile(loss='mse', optimizer=m_optimizer)

    print("----------Training-------------------")
    history_batch = []
    history_iter = []
    batch_amount = train_num // m_batchSize
    rest_size = train_num - (batch_amount*m_batchSize)

    # information of the best model on validation set.
    best_val_loss = 99999.99999
    best_model_name = "xxxxx"
    best_train_loss = 99999.99999

    for iter in range(m_epoch):

        # shuffle the training set
        index = [i for i in range(train_num)]
        np.random.shuffle(index)

        for bid in range(batch_amount):

            # using the shuffle index...
            x_batch = x_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]
            y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

            model.train_on_batch(x_batch, y_batch)    # training on batch

            if bid % m_print_loss_step == 0:
                logs = model.evaluate(x_batch, y_batch, verbose=0, batch_size=m_batchSize)
                print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                         GetProcessBar(bid, batch_amount),
                                                                         logs))
                history_batch.append([iter, bid, logs])

        # training on the rest data.
        if rest_size > 0:
            model.train_on_batch(x_train[index[batch_amount*m_batchSize : ]],
                                y_train[index[batch_amount*m_batchSize : ]])

        # one epoch is done. Do some information collections.
        train_iter_loss = model.evaluate(x_train, y_train, verbose=0, batch_size=m_batchSize)
        val_iter_loss = model.evaluate(x_val, y_val, verbose=0, batch_size=m_batchSize)
        print("#######>>>> iter({}/{}) train_loss={}, val_loss = {}".format(iter, m_epoch, train_iter_loss,
                                                                             val_iter_loss))
        history_iter.append([iter, train_iter_loss, val_iter_loss])

        if a.backup == True:   # whether to save the weight after each epoch
            model.save_weights(dir_results + "backup/model_{}_{}.h5".format(iter, val_iter_loss))

        if val_iter_loss < best_val_loss:  # save the best model on Validation set.
            RemoveDir(best_model_name)
            best_model_name = dir_results + "model_vggseg_onVal_{}.h5".format(val_iter_loss)
            model.save_weights(best_model_name)
            best_val_loss = val_iter_loss
            best_train_loss = train_iter_loss

    print("----------Training Done-------------------")
    # test on the testing set.
    model.load_weights(best_model_name)   # using the best model.

    test_loss = model.evaluate(x_test,y_test,verbose=0, batch_size=m_batchSize)

    # Save the predicted results and ground truth.
    MLAE_train = SavePredictedResult(x_train,y_train,'train')
    MLAE_val = SavePredictedResult(x_val, y_val, 'val')
    MLAE_test = SavePredictedResult(x_test,y_test,'test')

    # save the training information.
    wb = Workbook()
    ws1 = wb.active                         # MSE/MLAE
    ws1.title = "MLAE_MSE"
    ws2 = wb.create_sheet("EPOCH loss")      # iteration loss
    ws3 = wb.create_sheet("BATCH loss")     # batch loss

    ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
    ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])
    for i in range(len(history_iter)):
        ws2.append(history_iter[i])
    for i in range(len(history_batch)):
        ws3.append(history_batch[i])
    ws1.append(["Train loss", best_train_loss])
    ws1.append(["Val loss", best_val_loss])
    ws1.append(["Test loss", test_loss])
    ws1.append(["Train MLAE", MLAE_train])
    ws1.append(["val MLAE", MLAE_val])
    ws1.append(["Test MLAE", MLAE_test])

    wb.save(dir_results + "train_info.xlsx")

    print("Training MSE:", best_train_loss)
    print("Validat. MSE", best_val_loss)
    print("Testing MSE:", test_loss)
    print("Training MLAE:", MLAE_train)
    print("Validat. MLAE", MLAE_val)
    print("Testing MLAE:", MLAE_test)



