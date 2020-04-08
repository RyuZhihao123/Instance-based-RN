import numpy as np
import keras
import keras.backend as K
from keras.models import Sequential, Model
from keras.layers import Dense, Dropout, Flatten,Input
from keras.layers import Conv2D, MaxPooling2D
import tensorflow as tf
from keras.optimizers import SGD, Adam
import matplotlib.pyplot as plt
import cv2
import os
from  openpyxl import Workbook
import sklearn
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar,ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse
from utils.non_local import non_local_block


""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000


m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--savedir", default= 'RN_seg')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")
dir_results   = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # backup path. 'results/IRN_m'




def GetInformation(flag):
    if flag == 'train':
        return train_num, config.dir_subCharts_train, config.path_groundTruth_train
    if flag == 'val':
        return val_num, config.dir_subCharts_val, config.path_groundTruth_val
    if flag == 'test':
        return test_num, config.dir_subCharts_test, config.path_groundTruth_test

def LoadDataset(flag = 'train'):    # must be 'train' or 'val' or 'test'

    __max_load_num, __dir_subcharts, __path_groundTruth = GetInformation(flag)

    old_curpath = os.path.abspath(".")
    os.chdir(__dir_subcharts)    # change current path to load images faster.

    images1 = []  # The first  input image size =（max_load_num, h, w, 3)
    images2 = []  # The second input image size =（max_load_num, h, w, 3)
    labels = []   # groundTruth  size = (max_load_num, 1)

    for i in range(__max_load_num):

        images1.append(cv2.imread(__dir_subcharts+config.subChartName.format(i,0)))
        images2.append(cv2.imread(__dir_subcharts+config.subChartName.format(i,1)))

    lines = ReadLinesInFile(__path_groundTruth)
    for line in lines:
        elements = line.split()
        if len(elements) == 0:
            continue
        if len(labels) >= __max_load_num:
            break

        labels.append([float(p) for p in elements])


    images1 = np.array(images1, dtype= 'float32') /255.0
    images2 = np.array(images2, dtype= 'float32') /255.0
    labels = np.array(labels , dtype= 'float32')

    os.chdir(old_curpath)


    print("---------------------------")
    print("[Process] x1: ", images1.shape)
    print("[Process] x2: ", images2.shape)
    print("[Process] y: ", labels.shape)

    return images1, images2, labels

# IRN_m is final network to estimate the ratio vectors from multiple input instances.
def Build_RN_Network():
    # input layers.
    input_layers = []
    for i in range(2):
        input = Input(shape=(config.image_height, config.image_width, 3), name="input_{}".format(i))
        input_layers.append(input)

    # First extract individual features.
    individual_features = []
    for i in range(2):
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(input_layers[i])
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)
        x = Flatten()(x)
        individual_features.append(x)

    # g_theta
    g = []
    for i in range(len(individual_features)):
        for j in range(len(individual_features)):
            x = keras.layers.concatenate([individual_features[i], individual_features[j]])
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            g.append(x)
    f = keras.layers.average(g)

    # f_theta
    z = Dense(256, activation="relu")(f)
    z = Dense(256, activation="relu")(z)
    z = Dropout(0.5)(z)
    z = Dense(1, activation="linear")(z)

    return Model(inputs=input_layers, outputs=z)

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(x1,x2, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.predict([x1,x2], batch_size=1)
    predictFile = open(dir_results+flag+"_predicted_results.txt",'w')
    for i in range(x1.shape[0]):
        for t in range(dim):  # save the ground truth
            predictFile.write(str(y[i,t]) + '\t')
        for t in range(dim):  # save the predicted results
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':
    ClearDir(dir_results)
    ClearDir(dir_results+"backup")

    x1_train, x2_train, y_train = LoadDataset('train')
    x1_val, x2_val, y_val= LoadDataset('val')
    x1_test, x2_test, y_test = LoadDataset('test')

    ##  normalize to [-0.5, 0.5]
    x1_train -= .5
    x2_train -= .5
    x1_val -= .5
    x2_val -= .5
    x1_test -= .5
    x2_test -= .5


    print("----------Build Network----------------")
    model = Build_RN_Network()
    model.compile(loss='mse', optimizer=m_optimizer)

    print("----------Training-------------------")
    history_batch = []
    history_iter = []
    batch_amount = train_num // m_batchSize
    rest_size = train_num - (batch_amount*m_batchSize)

    # information of the best model on validation set.
    best_val_loss = 99999.99999
    best_model_name = "xxxxx"
    best_train_loss = 99999.99999


    for iter in range(m_epoch):

        # shuffle the training set !!!!!!!!!!!!!!
        index = [i for i in range(train_num)]
        np.random.shuffle(index)

        for bid in range(batch_amount):

            # using the shuffle index...
            x1_batch = x1_train[index[bid * m_batchSize : (bid + 1) * m_batchSize]]
            x2_batch = x2_train[index[bid * m_batchSize : (bid + 1) * m_batchSize]]
            y_batch = y_train[index[bid * m_batchSize : (bid + 1) * m_batchSize]]

            model.train_on_batch(x= [x1_batch,x2_batch], y=y_batch)  # training on batch

            if bid % m_print_loss_step == 0:
                logs = model.evaluate(x=[x1_batch,x2_batch],y=y_batch,verbose=0, batch_size=m_batchSize)
                print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                     GetProcessBar(bid, batch_amount), logs))
                history_batch.append([iter, bid, logs])

        # training on the rest data.
        if rest_size > 0:
            model.train_on_batch(x=[x1_train[index[batch_amount*m_batchSize : ]],
                                    x2_train[index[batch_amount*m_batchSize : ]]],
                                y=y_train[index[batch_amount*m_batchSize : ]])

        # one epoch is done. Do some information collections.
        train_iter_loss = model.evaluate([x1_train,x2_train], y_train, verbose=0, batch_size=m_batchSize)
        val_iter_loss = model.evaluate([x1_val, x2_val], y_val, verbose=0, batch_size=m_batchSize)
        print("#######>>>> epoch({}/{}) train_loss={}, val_loss = {}".format(iter, m_epoch, train_iter_loss,
                                                                                 val_iter_loss))
        history_iter.append([iter, train_iter_loss, val_iter_loss])

        if a.backup == True:  # whether to save the weight after each epoch
            model.save_weights(dir_results + "backup/model_{}_{}.h5".format(iter, val_iter_loss))

        if val_iter_loss < best_val_loss:  # save the best model on Validation set.
            RemoveDir(best_model_name)
            best_model_name = dir_results + "model_RNseg_onVal_{}.h5".format(val_iter_loss)
            model.save_weights(best_model_name)
            best_val_loss = val_iter_loss
            best_train_loss = train_iter_loss

    print("----------Training Done-------------------")
    # test on the testing set.
    model.load_weights(best_model_name)  # using the best model.

    test_loss = model.evaluate([x1_test, x2_test], y_test, verbose=0, batch_size=m_batchSize)

    # Save the predicted results,and return the MALE
    MLAE_train = SavePredictedResult(x1_train, x2_train, y_train, 'train')
    MLAE_val = SavePredictedResult(x1_val, x2_val, y_val, 'val')
    MLAE_test = SavePredictedResult(x1_test, x2_test, y_test, 'test')

    # save the training information.
    wb = Workbook()
    ws1 = wb.active  # MSE/MLAE
    ws1.title = "MLAE_MSE"
    ws2 = wb.create_sheet("EPOCH loss")  # iteration loss
    ws3 = wb.create_sheet("BATCH loss")  # batch loss

    ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
    ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])
    for i in range(len(history_iter)):
        ws2.append(history_iter[i])
    for i in range(len(history_batch)):
        ws3.append(history_batch[i])
    ws1.append(["Train loss", best_train_loss])
    ws1.append(["Val loss", best_val_loss])
    ws1.append(["Test loss", test_loss])
    ws1.append(["Train MLAE", MLAE_train])
    ws1.append(["val MLAE", MLAE_val])
    ws1.append(["Test MLAE", MLAE_test])

    wb.save(dir_results + "train_info.xlsx")

    print("Training MSE:", best_train_loss)
    print("Validat. MSE", best_val_loss)
    print("Testing MSE:", test_loss)
    print("Training MLAE:", MLAE_train)
    print("Validat. MLAE", MLAE_val)
    print("Testing MLAE:", MLAE_test)
