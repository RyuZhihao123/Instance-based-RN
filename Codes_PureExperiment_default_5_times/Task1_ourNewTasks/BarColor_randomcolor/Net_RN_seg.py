import numpy as np
import keras
import keras.backend as K
from keras.models import Sequential, Model
from keras.layers import Dense, Dropout, Flatten,Input
from keras.layers import Conv2D, MaxPooling2D
import tensorflow as tf
from keras.optimizers import SGD, Adam
import matplotlib.pyplot as plt
import cv2
import os
from  openpyxl import Workbook
import sklearn
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar,ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse
from utils.non_local import non_local_block
from Dataset_generator import GenerateDatasetVGG,GenerateDatasetIRNm
import pickle

""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000


m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--times", default=5, type=int)
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--savedir", default= 'RN_seg')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")

# IRN_m is final network to estimate the ratio vectors from multiple input instances.
def Build_RN_Network():
    # input layers.
    input_layers = []
    for i in range(config.max_obj_num):
        input = Input(shape=(config.image_height, config.image_width, 3), name="input_{}".format(i))
        input_layers.append(input)

    # First extract individual features.
    individual_features = []
    for i in range(config.max_obj_num):
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(input_layers[i])
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)

        x = Flatten()(x)

        x = Dense(24, activation='relu')(x)

        individual_features.append(x)

    # g_theta
    g = []
    for i in range(len(individual_features)):
        for j in range(len(individual_features)):
            x = keras.layers.concatenate([individual_features[i], individual_features[j]])
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            g.append(x)
    f = keras.layers.average(g)

    # f_theta
    z = Dense(256, activation="relu")(f)
    z = Dense(256, activation="relu")(z)
    z = Dropout(0.5)(z)
    z = Dense(config.max_obj_num, activation="linear")(z)

    return Model(inputs=input_layers, outputs=z)

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(dir_results, x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.predict(x=x, batch_size=1)
    predictFile = open(dir_results + flag + "_predicted_results.txt",'w')
    for i in range(y.shape[0]):
        for t in range(dim):  # save the ground_truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results.
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':
    dir_rootpath = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # ./results/network_name/
    ClearDir(dir_rootpath)

    exp_id = 0
    while exp_id < a.times:

        # current exp path
        dir_results = dir_rootpath+  "{}_{}/".format( a.savedir, exp_id )
        ClearDir(dir_results)
        ClearDir(dir_results + "backup")

        x_train, y_train = GenerateDatasetIRNm(flag='train', image_num=train_num)
        x_val, y_val = GenerateDatasetIRNm(flag='val', image_num=val_num)
        x_test, y_test = GenerateDatasetIRNm(flag='test', image_num=test_num)

        x_train -= .5
        x_val -= .5
        x_test -= .5

        # format the inputs as [img1, img2, img3,.....]
        x_train = [x_train[i] for i in range(config.max_obj_num)]
        x_val = [x_val[i] for i in range(config.max_obj_num)]
        x_test = [x_test[i] for i in range(config.max_obj_num)]

        print("----------Build Network----------------")
        model = Build_RN_Network()
        model.compile(loss='mse', optimizer=m_optimizer)

        print("----------Training-------------------")
        history_batch = []
        history_iter = []
        batch_amount = train_num // m_batchSize
        rest_size = train_num - (batch_amount * m_batchSize)

        # best model according to the training loss. (Since this network can't deal with this generalization task)
        best_train_loss = 99999.99999
        val_loss_using_Train = 99999.99999  # corresponding val loss using the best model on training set.
        best_model_name_onTrain = "xxxxx"

        # best model according to val loss.
        best_val_loss = 99999.999999
        best_model_name_onVal = "xxxxx"

        iter = 0
        while iter < m_epoch:

            # shuffle the training set !!!!!!!!!!!!!!
            index = [i for i in range(train_num)]
            np.random.shuffle(index)

            for bid in range(batch_amount):

                # using the shuffle index...
                x_batch = [x_train[i][index[bid * m_batchSize: (bid + 1) * m_batchSize]] for i in
                           range(config.max_obj_num)]
                y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

                model.train_on_batch(x=x_batch, y=y_batch)  # training on batch

                if bid % m_print_loss_step == 0:
                    logs = model.evaluate(x=x_batch, y=y_batch, verbose=0, batch_size=m_batchSize)
                    print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                             GetProcessBar(bid, batch_amount), logs))
                    history_batch.append([iter, bid, logs])

            # training on the rest data.
            if rest_size > 0:
                model.train_on_batch(
                    x=[x_train[i][index[batch_amount * m_batchSize:]] for i in range(config.max_obj_num)],
                    y=y_train[index[batch_amount * m_batchSize:]])

            # one epoch is done. Do some information collections.
            epoch_loss_train = model.evaluate(x=x_train, y=y_train, verbose=0, batch_size=m_batchSize)
            epoch_loss_val = model.evaluate(x=x_val, y=y_val, verbose=0, batch_size=m_batchSize)
            history_iter.append([iter, epoch_loss_train, epoch_loss_val])
            print(
                "----- epoch({}/{}) train_loss={}, val_loss={}".format(iter, m_epoch, epoch_loss_train, epoch_loss_val))

            # to avoid stuck in local optimum at the beginning
            iter += 1
            if iter >= 20 and epoch_loss_train > 0.05:
                history_iter.clear()
                history_batch.clear()
                best_train_loss = best_val_loss = val_loss_using_Train = 999999.

                model = Build_RN_Network()
                model.compile(loss='mse', optimizer=m_optimizer)
                iter = 0
                continue


            if epoch_loss_train < best_train_loss:
                RemoveDir(best_model_name_onTrain)
                best_model_name_onTrain = dir_results + "model_RNseg_onTrain_{}.h5".format(epoch_loss_val)
                model.save_weights(best_model_name_onTrain)
                val_loss_using_Train = epoch_loss_val
                best_train_loss = epoch_loss_train

            if epoch_loss_val < best_val_loss:  # save the best model on Validation set.
                best_val_loss = epoch_loss_val
                RemoveDir(best_model_name_onVal)
                best_model_name_onVal = dir_results + "model_RNseg_onVal_{}.h5".format(epoch_loss_val)
                model.save_weights(best_model_name_onVal)

        # test on the testing set.
        model.load_weights(best_model_name_onTrain)  # using the best model
        test_loss_usingTrain = model.evaluate(x_test, y_test, verbose=0, batch_size=m_batchSize)

        # Save the predicted results and return the MLAE.
        MLAE_train = SavePredictedResult(dir_results, x_train, y_train, 'train_usingTrain')
        MLAE_val = SavePredictedResult(dir_results, x_val, y_val, 'val_usingTrain')
        MLAE_test = SavePredictedResult(dir_results, x_test, y_test, 'test_usingTrain')

        model.load_weights(best_model_name_onVal)  # using the best model.
        train_loss_onVal = model.evaluate(x_train, y_train, verbose=0, batch_size=m_batchSize)
        test_loss_onVal = model.evaluate(x_test, y_test, verbose=0, batch_size=m_batchSize)
        MLAE_train_onVal = SavePredictedResult(dir_results, x_train, y_train, 'train_usingVal')
        MLAE_val_onVal = SavePredictedResult(dir_results, x_val, y_val, 'val_usingVal')
        MLAE_test_onVal = SavePredictedResult(dir_results, x_test, y_test, 'test_usingVal')

        # save the training information.
        wb = Workbook()
        ws1 = wb.active  # MSE/MLAE
        ws1.title = "MLAE_MSE"
        ws2 = wb.create_sheet("EPOCH loss")  # iteration loss
        ws3 = wb.create_sheet("BATCH loss")  # batch loss

        ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
        ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])

        for i in range(len(history_iter)):
            ws2.append(history_iter[i])
        for i in range(len(history_batch)):
            ws3.append(history_batch[i])

        ws1.append(["----------", "Using best model on train_set"])
        ws1.append(["Best Train loss", best_train_loss])
        ws1.append(["Val loss usingTrain", val_loss_using_Train])
        ws1.append(["Test loss usingTrain", test_loss_usingTrain])
        ws1.append(["Train MLAE", MLAE_train])
        ws1.append(["val MLAE", MLAE_val])
        ws1.append(["Test MLAE", MLAE_test])
        ws1.append(["----------", "Using best model on val_set    !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"])
        ws1.append(["Train loss usingVal", train_loss_onVal])
        ws1.append(["Best Val loss", best_val_loss])
        ws1.append(["Test loss usingVal", test_loss_onVal])
        ws1.append(["Train MLAE using Val", MLAE_train_onVal])
        ws1.append(["Val MLAE using Val", MLAE_val_onVal])
        ws1.append(["Test MLAE using Val", MLAE_test_onVal])

        wb.save(dir_results + "train_info.xlsx")

        print("-----Using the best model according to training loss-------")
        print("Training MSE:", best_train_loss)
        print("Validat. MSE", val_loss_using_Train)
        print("Testing MSE:", test_loss_usingTrain)
        print("Training MLAE:", MLAE_train)
        print("Validat. MLAE", MLAE_val)
        print("Testing MLAE:", MLAE_test)

        print("-----Using the best model according to Validation loss-------")
        print("Training MSE:", train_loss_onVal)
        print("Validat. MSE:", best_val_loss)
        print("Testing MSE:", test_loss_onVal)
        print("Training MLAE:", MLAE_train_onVal)
        print("Validat. MLAE", MLAE_val_onVal)
        print("Testing MLAE:", MLAE_test_onVal)

        ## save as pickle file
        stats = dict()
        stats_usingTrain = dict()
        stats_usingTrain['MSE_train'] = best_train_loss
        stats_usingTrain['MSE_val'] = val_loss_using_Train
        stats_usingTrain['MSE_test'] = test_loss_usingTrain
        stats_usingTrain['MLAE_train'] = MLAE_train
        stats_usingTrain['MLAE_val'] = MLAE_val
        stats_usingTrain['MLAE_test'] = MLAE_test

        stats_usingVal = dict()
        stats_usingVal['MSE_train'] = train_loss_onVal
        stats_usingVal['MSE_val'] = best_val_loss
        stats_usingVal['MSE_test'] = test_loss_onVal
        stats_usingVal['MLAE_train'] = MLAE_train_onVal
        stats_usingVal['MLAE_val'] = MLAE_val_onVal
        stats_usingVal['MLAE_test'] = MLAE_test_onVal

        stats['Results_using_BestModel_OnValSet'] = stats_usingVal
        stats['Results_using_BestModel_OnTrainSet'] = stats_usingTrain
        stats['loss_train'] = [history_iter[i][1] for i in range(len(history_iter))]
        stats['loss_val'] = [history_iter[i][2] for i in range(len(history_iter))]

        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'wb') as f:
            pickle.dump(stats, f)
            f.close()

        exp_id += 1

    # compute average MLAE, MSE and SD.
    MLAE_trains_onTrain = []
    MLAE_tests_onTrain = []
    MSE_trains_onTrain = []
    MSE_tests_onTrain = []

    MLAE_trains_onVal = []
    MLAE_tests_onVal = []
    MSE_trains_onVal = []
    MSE_tests_onVal = []

    for exp_id in range(a.times):
        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'rb') as f:
            stats = pickle.load(f)

            MLAE_trains_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MLAE_train'])
            MLAE_tests_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MLAE_test'])
            MSE_trains_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MSE_train'])
            MSE_tests_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MSE_test'])

            MLAE_trains_onVal.append(stats['Results_using_BestModel_OnValSet']['MLAE_train'])
            MLAE_tests_onVal.append(stats['Results_using_BestModel_OnValSet']['MLAE_test'])
            MSE_trains_onVal.append(stats['Results_using_BestModel_OnValSet']['MSE_train'])
            MSE_tests_onVal.append(stats['Results_using_BestModel_OnValSet']['MSE_test'])

            f.close()

    with open(dir_rootpath + "{}_avg.p".format(a.savedir), 'wb') as f:
        stats = dict()
        stats_usingTrain = dict()
        stats_usingTrain['MSE_train_avg'] = np.average(MSE_trains_onTrain)
        stats_usingTrain['MSE_test_avg'] = np.average(MSE_tests_onTrain)
        stats_usingTrain['MSE_train_SD'] = np.std(MSE_trains_onTrain)
        stats_usingTrain['MSE_test_SD'] = np.std(MSE_tests_onTrain)
        stats_usingTrain['MLAE_train_avg'] = np.average(MLAE_trains_onTrain)
        stats_usingTrain['MLAE_test_avg'] = np.average(MLAE_tests_onTrain)
        stats_usingTrain['MLAE_train_SD'] = np.std(MLAE_trains_onTrain)
        stats_usingTrain['MLAE_test_SD'] = np.std(MLAE_tests_onTrain)

        stats_usingVal = dict()
        stats_usingVal['MSE_train_avg'] = np.average(MSE_trains_onVal)
        stats_usingVal['MSE_test_avg'] = np.average(MSE_tests_onVal)
        stats_usingVal['MSE_train_SD'] = np.std(MSE_trains_onVal)
        stats_usingVal['MSE_test_SD'] = np.std(MSE_tests_onVal)

        stats_usingVal['MLAE_train_avg'] = np.average(MLAE_trains_onVal)
        stats_usingVal['MLAE_test_avg'] = np.average(MLAE_tests_onVal)
        stats_usingVal['MLAE_train_SD'] = np.std(MLAE_trains_onVal)
        stats_usingVal['MLAE_test_SD'] = np.std(MLAE_tests_onVal)

        stats['Results_using_BestModel_OnValSet'] = stats_usingVal
        stats['Results_using_BestModel_OnTrainSet'] = stats_usingTrain
        pickle.dump(stats, f)

        f.close()