import numpy as np
import tensorflow as tf
from keras.optimizers import SGD,Adam
import matplotlib.pyplot as plt
from  openpyxl import  Workbook
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar, ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse, os, cv2, keras, sklearn
from utils.model_rn import ModelRN
from Dataset_generator import GenerateDatasetVGG,GenerateDatasetIRNm
import pickle
from keras import backend as K


""" some important configuration """
train_num = 60000            # image number.
val_num   = 20000
test_num  = 20000

m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--times", default=5, type=int)
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--savedir", default= 'RN')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")


# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(dir_results, sess, x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.GetPredictions(sess,x,y,x.shape[0])

    predictFile = open(dir_results+flag+"_pred_results.txt",'w')
    for i in range(predict_Y.shape[0]):
        for t in range(dim):  # save the ground truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y[:predict_Y.shape[0]] * 100) + .125)

    return MLAE, y[:predict_Y.shape[0]], predict_Y

if __name__ == '__main__':

    dir_rootpath = os.path.abspath(".") + "/results/{}/".format(a.savedir)  # ./results/network_name/
    ClearDir(dir_rootpath)

    exp_id = 0
    while exp_id < a.times:

        # current exp path
        dir_results = dir_rootpath+  "{}_{}/".format( a.savedir, exp_id  )  # backup path. 'results/VGG'
        ClearDir(dir_results)

        ClearDir(dir_results + "best_model_onTrainset")
        ClearDir(dir_results + "best_model_onValset")

        x_train, y_train = GenerateDatasetVGG('train', train_num)
        x_val, y_val = GenerateDatasetVGG('val', val_num)
        x_test, y_test = GenerateDatasetVGG('test', test_num)

        x_train -= .5
        x_val -= .5
        x_test -= .5

        print("----------Build network----------------")

        tf.reset_default_graph()
        model = ModelRN(learning_rate=a.lr, batch_size=m_batchSize, c_dim=3, a_dim=config.max_obj_num)

        # TensorFlow environment.
        init = tf.global_variables_initializer()
        sess = tf.Session()
        sess.run(init)

        saver = tf.train.Saver(max_to_keep=1)
        saver_val = tf.train.Saver(max_to_keep=1)

        print("----------Training-------------------")
        history_batch = []
        history_iter = []
        batch_amount = train_num // m_batchSize
        rest_size = train_num - (batch_amount * m_batchSize)

        # best model according to the training loss. (Since this network can't deal with this generalization task)
        best_train_loss = 99999.99999
        val_loss_using_Train = 99999.99999  # corresponding val loss using the best model on training set.
        best_model_name_onTrain = "xxxxx"

        # best model according to val loss.
        best_val_loss = 99999.999999
        best_model_name_onVal = "xxxxx"

        iter = 0
        while iter < m_epoch:

            # shuffle the training set
            index = [i for i in range(train_num)]
            np.random.shuffle(index)

            for bid in range(batch_amount):

                # using the shuffle index...
                x_batch = x_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]
                y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

                loss = model.Run_one_batch(sess, x_batch, y_batch)

                if bid % m_print_loss_step == 0:
                    print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                             GetProcessBar(bid, batch_amount), loss))
                    history_batch.append([iter, bid, loss])

            # one epoch is done. Do some information collections.
            train_iter_loss = model.GetTotalLoss(sess, x_train, y_train, x_train.shape[0])
            val_iter_loss = model.GetTotalLoss(sess, x_val, y_val, x_val.shape[0])
            print("#######>>>> iter({}/{}) train_loss={}, val_loss = {}".format(iter, m_epoch, train_iter_loss,
                                                                                val_iter_loss))
            history_iter.append([iter, train_iter_loss, val_iter_loss])


            iter += 1

            # For generalization task, we save both the best models on training set and on validation set.
            # since this network couldn't get good result on validation sets.
            if train_iter_loss < best_train_loss:
                RemoveDir(best_model_name_onTrain)
                best_model_name_onTrain = dir_results + "best_model_onTrainset/model_RN_onTrain_{}.ckpt".format(
                    val_iter_loss)
                saver.save(sess, best_model_name_onTrain)
                val_loss_using_Train = val_iter_loss
                best_train_loss = train_iter_loss

            if val_iter_loss < best_val_loss:
                best_val_loss = val_iter_loss
                RemoveDir(best_model_name_onVal)
                best_model_name_onVal = dir_results + "best_model_onValset/model_RN_onVal_{}.ckpt".format(val_iter_loss)
                saver_val.save(sess, best_model_name_onVal)

        print("----------Training Done-------------------")
        saver.restore(sess, best_model_name_onTrain)

        # test on the testing set.
        test_loss_usingTrain = model.GetTotalLoss(sess, x_test, y_test, x_test.shape[0])

        # Save the predicted results and ground truth.
        MLAE_train,_,_ = SavePredictedResult(dir_results, sess, x_train, y_train, 'trainset_usingTrain')
        MLAE_val,_,_ = SavePredictedResult(dir_results, sess, x_val, y_val, 'valset_usingTrain')
        MLAE_test,_,_ = SavePredictedResult(dir_results, sess, x_test, y_test, 'testset_usingTrain')

        saver_val.restore(sess, best_model_name_onVal)
        train_loss_onVal = model.GetTotalLoss(sess, x_train, y_train, x_train.shape[0])
        test_loss_onVal = model.GetTotalLoss(sess, x_test, y_test, x_test.shape[0])
        MLAE_train_onVal,_,_ = SavePredictedResult(dir_results, sess, x_train, y_train, 'trainset_usingVal')
        MLAE_val_onVal,_,_ = SavePredictedResult(dir_results, sess, x_val, y_val, 'valset_usingVal')
        MLAE_test_onVal, _y_test, _y_pred = SavePredictedResult(dir_results, sess, x_test, y_test, 'testset_usingVal')

        # save the training information.
        wb = Workbook()
        ws1 = wb.active  # MSE/MLAE
        ws1.title = "MLAE_MSE"
        ws2 = wb.create_sheet("EPOCH loss")  # iteration loss
        ws3 = wb.create_sheet("BATCH loss")  # batch loss

        ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
        ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])

        for i in range(len(history_iter)):
            ws2.append(history_iter[i])
        for i in range(len(history_batch)):
            ws3.append(history_batch[i])
        ws1.append(["----------", "Using best model on train_set    !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"])
        ws1.append(["Best Train loss", best_train_loss])
        ws1.append(["Val loss usingTrain", val_loss_using_Train])
        ws1.append(["Test loss usingTrain", test_loss_usingTrain])
        ws1.append(["Train MLAE", MLAE_train])
        ws1.append(["val MLAE", MLAE_val])
        ws1.append(["Test MLAE", MLAE_test])
        ws1.append(["----------", "Using best model on val_set"])
        ws1.append(["Train loss usingVal", train_loss_onVal])
        ws1.append(["Best Val loss", best_val_loss])
        ws1.append(["Test loss usingVal", test_loss_onVal])
        ws1.append(["Train MLAE using Val", MLAE_train_onVal])
        ws1.append(["Val MLAE using Val", MLAE_val_onVal])
        ws1.append(["Test MLAE using Val", MLAE_test_onVal])

        wb.save(dir_results + "train_info.xlsx")

        print("-----Using the best model on training loss-------")
        print("Training MSE:", best_train_loss)
        print("Validat. MSE", val_loss_using_Train)
        print("Testing MSE:", test_loss_usingTrain)
        print("Training MLAE:", MLAE_train)
        print("Validat. MLAE", MLAE_val)
        print("Testing MLAE:", MLAE_test)
        print("-----Using the best model on Validation loss-------")
        print("Training MSE:", train_loss_onVal)
        print("Validat. MSE:", best_val_loss)
        print("Testing MSE:", test_loss_onVal)
        print("Training MLAE:", MLAE_train_onVal)
        print("Validat. MLAE", MLAE_val_onVal)
        print("Testing MLAE:", MLAE_test_onVal)

        sess.close()
        ## save as pickle file
        stats = dict()
        stats_usingTrain = dict()
        stats_usingTrain['MSE_train'] = best_train_loss
        stats_usingTrain['MSE_val'] = val_loss_using_Train
        stats_usingTrain['MSE_test'] = test_loss_usingTrain
        stats_usingTrain['MLAE_train'] = MLAE_train
        stats_usingTrain['MLAE_val'] = MLAE_val
        stats_usingTrain['MLAE_test'] = MLAE_test

        stats_usingVal = dict()
        stats_usingVal['MSE_train'] = train_loss_onVal
        stats_usingVal['MSE_val'] = best_val_loss
        stats_usingVal['MSE_test'] = test_loss_onVal
        stats_usingVal['MLAE_train'] = MLAE_train_onVal
        stats_usingVal['MLAE_val'] = MLAE_val_onVal
        stats_usingVal['MLAE_test'] = MLAE_test_onVal

        stats['Results_using_BestModel_OnValSet'] = stats_usingVal
        stats['Results_using_BestModel_OnTrainSet'] = stats_usingTrain
        stats['loss_train'] = [history_iter[i][1] for i in range(len(history_iter))]
        stats['loss_val'] = [history_iter[i][2] for i in range(len(history_iter))]

        stats['y_test'] = _y_test
        stats['y_pred'] = _y_pred

        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'wb') as f:
            pickle.dump(stats, f)
            f.close()

        exp_id += 1

    # compute average MLAE, MSE and SD.
    MLAE_trains_onTrain = []
    MLAE_tests_onTrain = []
    MSE_trains_onTrain = []
    MSE_tests_onTrain = []

    MLAE_trains_onVal = []
    MLAE_tests_onVal = []
    MSE_trains_onVal = []
    MSE_tests_onVal = []

    for exp_id in range(a.times):
        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'rb') as f:
            stats = pickle.load(f)

            MLAE_trains_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MLAE_train'])
            MLAE_tests_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MLAE_test'])
            MSE_trains_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MSE_train'])
            MSE_tests_onTrain.append(stats['Results_using_BestModel_OnTrainSet']['MSE_test'])

            MLAE_trains_onVal.append(stats['Results_using_BestModel_OnValSet']['MLAE_train'])
            MLAE_tests_onVal.append(stats['Results_using_BestModel_OnValSet']['MLAE_test'])
            MSE_trains_onVal.append(stats['Results_using_BestModel_OnValSet']['MSE_train'])
            MSE_tests_onVal.append(stats['Results_using_BestModel_OnValSet']['MSE_test'])

            f.close()

    with open(dir_rootpath + "{}_avg.p".format(a.savedir), 'wb') as f:
        stats = dict()
        stats_usingTrain = dict()
        stats_usingTrain['MSE_train_avg'] = np.average(MSE_trains_onTrain)
        stats_usingTrain['MSE_test_avg'] = np.average(MSE_tests_onTrain)
        stats_usingTrain['MSE_train_SD'] = np.std(MSE_trains_onTrain)
        stats_usingTrain['MSE_test_SD'] = np.std(MSE_tests_onTrain)
        stats_usingTrain['MLAE_train_avg'] = np.average(MLAE_trains_onTrain)
        stats_usingTrain['MLAE_test_avg'] = np.average(MLAE_tests_onTrain)
        stats_usingTrain['MLAE_train_SD'] = np.std(MLAE_trains_onTrain)
        stats_usingTrain['MLAE_test_SD'] = np.std(MLAE_tests_onTrain)

        stats_usingVal = dict()
        stats_usingVal['MSE_train_avg'] = np.average(MSE_trains_onVal)
        stats_usingVal['MSE_test_avg'] = np.average(MSE_tests_onVal)
        stats_usingVal['MSE_train_SD'] = np.std(MSE_trains_onVal)
        stats_usingVal['MSE_test_SD'] = np.std(MSE_tests_onVal)

        stats_usingVal['MLAE_train_avg'] = np.average(MLAE_trains_onVal)
        stats_usingVal['MLAE_test_avg'] = np.average(MLAE_tests_onVal)
        stats_usingVal['MLAE_train_SD'] = np.std(MLAE_trains_onVal)
        stats_usingVal['MLAE_test_SD'] = np.std(MLAE_tests_onVal)

        stats['Results_using_BestModel_OnValSet'] = stats_usingVal
        stats['Results_using_BestModel_OnTrainSet'] = stats_usingTrain
        pickle.dump(stats, f)

        f.close()