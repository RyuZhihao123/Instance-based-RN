import numpy as np
import keras
import keras.backend as K
from keras.models import Sequential, Model
from keras.layers import Dense, Dropout, Flatten,Input
from keras.layers import Conv2D, MaxPooling2D
import tensorflow as tf
from keras.optimizers import SGD, Adam
import matplotlib.pyplot as plt
import cv2
import os
import sklearn
from  openpyxl import Workbook
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar,ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse, copy
from utils.non_local import non_local_block
from Dataset_generator import GenerateDatasetVGG,GenerateDatasetIRN
import pickle

""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000

m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--times", default=5, type=int)
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--savedir", default= 'IRNp')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")


# Level1 module is to extract the individual features from one instance.
# Level1 has two NON-LOCAL block.
def Level1_Module():
    input = Input(shape=(config.image_height, config.image_width, 1))
    x = Conv2D(32, (3, 3), activation='relu', padding='same')(input)
    x = Conv2D(32, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = non_local_block(x)    # non local block
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = non_local_block(x)    # non local block
    return Model(inputs=input, outputs=x)

# IRN_p module is to estimate the ratio between one pair.
def Build_IRN_p_Network():
    inputA = Input(shape=(config.image_height, config.image_width, 1))
    inputB = Input(shape=(config.image_height, config.image_width, 1))

    level1 =Level1_Module()    # build a level1 module
    a = level1(inputA)         # extract two individual features.
    b = level1(inputB)

    combined = keras.layers.concatenate([a, b])   # concatenate them.
    z = Conv2D(64, (3, 3), activation='relu',padding='same')(combined)
    z = Conv2D(64, (3, 3), activation='relu', padding='same')(z)
    z = non_local_block(z)        # non local block
    #
    z = Flatten()(z)
    z = Dense(256, activation="relu")(z)
    z = Dropout(0.5)(z)
    z = Dense(1, activation="linear")(z)    # output the ratio of this pair.

    return Model(inputs=[inputA, inputB], outputs=z)

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(dir_results, x1,x2, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.predict([x1,x2], batch_size=1)
    predictFile = open(dir_results+flag+"_predicted_results.txt",'w')
    for i in range(x1.shape[0]):
        for t in range(dim):  # save the ground truth
            predictFile.write(str(y[i,t]) + '\t')
        for t in range(dim):  # save the predicted results
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':

    dir_rootpath = os.path.abspath(".") + "/results/{}/".format(a.savedir)  # ./results/network_name/
    ClearDir(dir_rootpath)

    exp_id = 0
    while exp_id < a.times:

        # current exp path
        dir_results = dir_rootpath+  "{}_{}/".format( a.savedir, exp_id)  # backup path. 'results/VGG'
        ClearDir(dir_results)
        ClearDir(dir_results + "backup")

        x1_train, x2_train, y_train = GenerateDatasetIRN('train', train_num)
        x1_val, x2_val, y_val = GenerateDatasetIRN('val', val_num)
        x1_test, x2_test, y_test = GenerateDatasetIRN('test', test_num)

        ##  normalize to [-0.5, 0.5]
        x1_train -= .5
        x2_train -= .5
        x1_val -= .5
        x2_val -= .5
        x1_test -= .5
        x2_test -= .5

        print("----------Build Network----------------")
        model = Build_IRN_p_Network()
        model.compile(optimizer=m_optimizer, loss='mse')

        print("----------Training-------------------")
        history_batch = []
        history_iter = []
        batch_amount = train_num // m_batchSize
        rest_size = train_num - (batch_amount * m_batchSize)

        # information of the best model on validation set.
        best_val_loss = 99999.99999
        best_model_name = "xxxxx"
        best_train_loss = 99999.99999

        iter = 0
        while iter < m_epoch:

            # shuffle the training set
            index = [i for i in range(train_num)]
            np.random.shuffle(index)

            for bid in range(batch_amount):

                # using the shuffle index...
                x1_batch = x1_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]
                x2_batch = x2_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]
                y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

                model.train_on_batch(x=[x1_batch, x2_batch], y=y_batch)  # training on batch

                if bid % m_print_loss_step == 0:
                    logs = model.evaluate(x=[x1_batch, x2_batch], y=y_batch, verbose=0, batch_size=m_batchSize)
                    print("iter({}/{}) Batch({}/{}) {} : MSE_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                             GetProcessBar(bid, batch_amount), logs))
                    history_batch.append([iter, bid, logs])

            # training on the rest data.
            if rest_size > 0:
                model.train_on_batch(x=[x1_train[index[batch_amount * m_batchSize:]],
                                        x2_train[index[batch_amount * m_batchSize:]]],
                                     y=y_train[index[batch_amount * m_batchSize:]])

            # one epoch is done. Do some information collections.
            train_iter_loss = model.evaluate([x1_train, x2_train], y_train, verbose=0, batch_size=m_batchSize)
            val_iter_loss = model.evaluate([x1_val, x2_val], y_val, verbose=0, batch_size=m_batchSize)
            print("#######>>>> epoch({}/{}) train_loss={}, val_loss = {}".format(iter, m_epoch, train_iter_loss,
                                                                                 val_iter_loss))
            history_iter.append([iter, train_iter_loss, val_iter_loss])

            # to avoid stuck in local optimum at the beginning
            iter += 1
            if iter >= 10 and iter < 15 and best_train_loss > 0.05:
                history_iter.clear()
                history_batch.clear()
                best_train_loss = best_val_loss = 999999.

                model = Build_IRN_p_Network()  # reset the network.
                model.compile(loss='mse', optimizer=m_optimizer)
                iter = 0
                continue

            if val_iter_loss < best_val_loss:  # save the best model on Validation set.
                RemoveDir(best_model_name)
                best_model_name = dir_results + "model_IRNp_onVal_{}.h5".format(val_iter_loss)
                model.save_weights(best_model_name)
                best_val_loss = val_iter_loss
                best_train_loss = train_iter_loss

        print("----------Training Done-------------------")
        # test on the testing set.
        model.load_weights(best_model_name)  # using the best model.

        test_loss = model.evaluate([x1_test, x2_test], y_test, verbose=0, batch_size=m_batchSize)

        # Save the predicted results,and return the MALE
        MLAE_train = SavePredictedResult(dir_results, x1_train, x2_train, y_train, 'train')
        MLAE_val = SavePredictedResult(dir_results, x1_val, x2_val, y_val, 'val')
        MLAE_test = SavePredictedResult(dir_results, x1_test, x2_test, y_test, 'test')

        # save the training information.
        wb = Workbook()
        ws1 = wb.active  # MSE/MLAE
        ws1.title = "MLAE_MSE"
        ws2 = wb.create_sheet("EPOCH loss")  # iteration loss
        ws3 = wb.create_sheet("BATCH loss")  # batch loss

        ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
        ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])
        for i in range(len(history_iter)):
            ws2.append(history_iter[i])
        for i in range(len(history_batch)):
            ws3.append(history_batch[i])
        ws1.append(["Train loss", best_train_loss])
        ws1.append(["Val loss", best_val_loss])
        ws1.append(["Test loss", test_loss])
        ws1.append(["Train MLAE", MLAE_train])
        ws1.append(["val MLAE", MLAE_val])
        ws1.append(["Test MLAE", MLAE_test])

        wb.save(dir_results + "train_info.xlsx")

        print("Training MSE:", best_train_loss)
        print("Validat. MSE", best_val_loss)
        print("Testing MSE:", test_loss)
        print("Training MLAE:", MLAE_train)
        print("Validat. MLAE", MLAE_val)
        print("Testing MLAE:", MLAE_test)

        ## save as pickle file
        stats = dict()

        stats['MSE_train'] = best_train_loss
        stats['MSE_val'] = best_val_loss
        stats['MSE_test'] = test_loss

        stats['MLAE_train'] = MLAE_train
        stats['MLAE_val'] = MLAE_val
        stats['MLAE_test'] = MLAE_test

        stats['loss_train'] = [history_iter[i][1] for i in range(len(history_iter))]
        stats['loss_val'] = [history_iter[i][2] for i in range(len(history_iter))]

        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'wb') as f:
            pickle.dump(stats, f)
            f.close()

        exp_id += 1

    # compute average MLAE, MSE and SD.
    MLAE_trains = []
    MLAE_tests = []

    MSE_trains = []
    MSE_tests = []

    for exp_id in range(a.times):
        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'rb') as f:
            stats = pickle.load(f)

            MLAE_trains.append(stats['MLAE_train'])
            MLAE_tests.append(stats['MLAE_test'])
            MSE_trains.append(stats['MSE_train'])
            MSE_tests.append(stats['MSE_test'])

            f.close()

    with open(dir_rootpath + "{}_avg.p".format(a.savedir), 'wb') as f:
        stats = dict()

        stats['MSE_train_avg'] = np.average(MSE_trains)
        stats['MSE_test_avg'] = np.average(MSE_tests)
        stats['MSE_train_SD'] = np.std(MSE_trains)
        stats['MSE_test_SD'] = np.std(MSE_tests)

        stats['MLAE_train_avg'] = np.average(MLAE_trains)
        stats['MLAE_test_avg'] = np.average(MLAE_tests)
        stats['MLAE_train_SD'] = np.std(MLAE_trains)
        stats['MLAE_test_SD'] = np.std(MLAE_tests)
        pickle.dump(stats, f)

        f.close()

