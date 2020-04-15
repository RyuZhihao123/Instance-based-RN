import numpy as np
import keras
import keras.backend as K
from keras.models import Sequential, Model
from keras.layers import Dense, Dropout, Flatten,Input
from keras.layers import Conv2D, MaxPooling2D
import tensorflow as tf
from keras.optimizers import SGD, Adam
import matplotlib.pyplot as plt
import cv2
import os
from  openpyxl import Workbook
import sklearn
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar,ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse
from utils.non_local import non_local_block
from Dataset_generator import GenerateDatasetVGG,GenerateDatasetIRNm
import pickle



""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000


m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--times", default=5, type=int)
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--savedir", default= 'RN_seg')         # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")

# IRN_m is final network to estimate the ratio vectors from multiple input instances.
def Build_RN_Network():
    # input layers.
    input_layers = []
    for i in range(config.max_obj_num):
        input = Input(shape=(config.image_height, config.image_width, 3), name="input_{}".format(i))
        input_layers.append(input)

    # First extract individual features.
    individual_features = []
    for i in range(config.max_obj_num):
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(input_layers[i])
        x = Conv2D(24,(5,5),strides=(3,3),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)
        x = Conv2D(24,(5,5),strides=(2,2),activation='relu',padding='same')(x)

        x = Flatten()(x)

        x = Dense(24, activation='relu')(x)

        individual_features.append(x)

    # g_theta
    g = []
    for i in range(len(individual_features)):
        for j in range(len(individual_features)):
            x = keras.layers.concatenate([individual_features[i], individual_features[j]])
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            x = Dense(256, activation="relu")(x)
            g.append(x)
    f = keras.layers.average(g)

    # f_theta
    z = Dense(256, activation="relu")(f)
    z = Dense(256, activation="relu")(z)
    z = Dropout(0.5)(z)
    z = Dense(config.max_obj_num, activation="linear")(z)

    return Model(inputs=input_layers, outputs=z)

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(dir_results, x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    predict_Y = model.predict(x=x, batch_size=1)
    predictFile = open(dir_results + flag + "_predicted_results.txt",'w')
    for i in range(y.shape[0]):
        for t in range(dim):  # save the ground_truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results.
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':
    dir_rootpath = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # ./results/network_name/
    ClearDir(dir_rootpath)

    exp_id = 0
    while exp_id < a.times:

        # current exp path
        dir_results = dir_rootpath+  "{}_{}/".format( a.savedir, exp_id )
        ClearDir(dir_results)
        ClearDir(dir_results + "backup")

        x_train, y_train = GenerateDatasetIRNm(flag='train', image_num=train_num)
        x_val, y_val = GenerateDatasetIRNm(flag='val', image_num=val_num)
        x_test, y_test = GenerateDatasetIRNm(flag='test', image_num=test_num)

        x_train -= .5
        x_val -= .5
        x_test -= .5

        # format the inputs as [img1, img2, img3,.....]
        x_train = [x_train[i] for i in range(config.max_obj_num)]
        x_val = [x_val[i] for i in range(config.max_obj_num)]
        x_test = [x_test[i] for i in range(config.max_obj_num)]

        print("----------Build Network----------------")
        model = Build_RN_Network()
        model.compile(loss='mse', optimizer=m_optimizer)

        print("----------Training-------------------")
        history_batch = []
        history_iter = []
        batch_amount = train_num // m_batchSize
        rest_size = train_num - (batch_amount * m_batchSize)
        # information of the best model on validation set.
        best_val_loss = 99999.99999
        best_model_name = "xxxxx"
        best_train_loss = 99999.99999

        iter = 0
        while iter < m_epoch:

            # shuffle the training set !!!!!!!!!!!!!!
            index = [i for i in range(train_num)]
            np.random.shuffle(index)

            for bid in range(batch_amount):

                # using the shuffle index...
                x_batch = [x_train[i][index[bid * m_batchSize: (bid + 1) * m_batchSize]] for i in
                           range(config.max_obj_num)]
                y_batch = y_train[index[bid * m_batchSize: (bid + 1) * m_batchSize]]

                model.train_on_batch(x=x_batch, y=y_batch)  # training on batch

                if bid % m_print_loss_step == 0:
                    logs = model.evaluate(x=x_batch, y=y_batch, verbose=0, batch_size=m_batchSize)
                    print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                             GetProcessBar(bid, batch_amount), logs))
                    history_batch.append([iter, bid, logs])

            # training on the rest data.
            if rest_size > 0:
                model.train_on_batch(
                    x=[x_train[i][index[batch_amount * m_batchSize:]] for i in range(config.max_obj_num)],
                    y=y_train[index[batch_amount * m_batchSize:]])

            # one epoch is done. Do some information collections.
            epoch_loss_train = model.evaluate(x=x_train, y=y_train, verbose=0, batch_size=m_batchSize)
            epoch_loss_val = model.evaluate(x=x_val, y=y_val, verbose=0, batch_size=m_batchSize)
            history_iter.append([iter, epoch_loss_train, epoch_loss_val])
            print("----- epoch({}/{}) train_loss={}, val_loss={}".format(iter, m_epoch, epoch_loss_train, epoch_loss_val))

            # to avoid stuck in local optimum at the beginning
            iter += 1
            if iter >= 20 and epoch_loss_train > 0.03:
                history_iter.clear()
                history_batch.clear()
                best_train_loss = best_val_loss = 999999.

                model = Build_RN_Network()
                model.compile(loss='mse', optimizer=m_optimizer)
                iter = 0
                continue

            if epoch_loss_val < best_val_loss:  # save the best model on Validation set.
                RemoveDir(best_model_name)
                best_model_name = dir_results + "model_RNseg_{}.h5".format(epoch_loss_val)
                model.save_weights(best_model_name)
                best_val_loss = epoch_loss_val
                best_train_loss = epoch_loss_train

        # test on the testing set.
        model.load_weights(best_model_name)  # using the best model
        test_loss = model.evaluate(x_test, y_test, verbose=0, batch_size=m_batchSize)

        # Save the predicted results and return the MLAE.
        MLAE_train = SavePredictedResult(dir_results, x_train, y_train, 'train')
        MLAE_val = SavePredictedResult(dir_results, x_val, y_val, 'val')
        MLAE_test = SavePredictedResult(dir_results, x_test, y_test, 'test')

        # save the training information.
        wb = Workbook()
        ws1 = wb.active  # MSE/MLAE
        ws1.title = "MLAE_MSE"
        ws2 = wb.create_sheet("EPOCH loss")  # iteration loss
        ws3 = wb.create_sheet("BATCH loss")  # batch loss

        ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
        ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])

        for i in range(len(history_iter)):
            ws2.append(history_iter[i])
        for i in range(len(history_batch)):
            ws3.append(history_batch[i])

        ws1.append(["Train loss", best_train_loss])
        ws1.append(["Val loss", best_val_loss])
        ws1.append(["Test loss", test_loss])
        ws1.append(["Train MLAE", MLAE_train])
        ws1.append(["val MLAE", MLAE_val])
        ws1.append(["Test MLAE", MLAE_test])

        wb.save(dir_results + "train_info.xlsx")

        print("Training MSE:", best_train_loss)
        print("Validat. MSE:", best_val_loss)
        print("Testing MSE:", test_loss)
        print("Training MLAE:", MLAE_train)
        print("Validat. MLAE:", MLAE_val)
        print("Testing MLAE:", MLAE_test)

        ## save as pickle file

        stats = dict()

        stats['MSE_train'] = best_train_loss
        stats['MSE_val'] = best_val_loss
        stats['MSE_test'] = test_loss

        stats['MLAE_train'] = MLAE_train
        stats['MLAE_val'] = MLAE_val
        stats['MLAE_test'] = MLAE_test

        stats['loss_train'] = [history_iter[i][1] for i in range(len(history_iter))]
        stats['loss_val'] = [history_iter[i][2] for i in range(len(history_iter))]

        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'wb') as f:
            pickle.dump(stats, f)
            f.close()

        exp_id += 1

        # compute average MLAE, MSE and SD.
    MLAE_trains = []
    MLAE_tests = []

    MSE_trains = []
    MSE_tests = []

    for exp_id in range(a.times):
        with open(dir_rootpath + "{}_{}.p".format(a.savedir, exp_id), 'rb') as f:
            stats = pickle.load(f)

            MLAE_trains.append(stats['MLAE_train'])
            MLAE_tests.append(stats['MLAE_test'])
            MSE_trains.append(stats['MSE_train'])
            MSE_tests.append(stats['MSE_test'])

            f.close()

    with open(dir_rootpath + "{}_avg.p".format(a.savedir), 'wb') as f:
        stats = dict()

        stats['MSE_train_avg'] = np.average(MSE_trains)
        stats['MSE_test_avg'] = np.average(MSE_tests)
        stats['MSE_train_SD'] = np.std(MSE_trains)
        stats['MSE_test_SD'] = np.std(MSE_tests)

        stats['MLAE_train_avg'] = np.average(MLAE_trains)
        stats['MLAE_test_avg'] = np.average(MLAE_tests)
        stats['MLAE_train_SD'] = np.std(MLAE_trains)
        stats['MLAE_test_SD'] = np.std(MLAE_tests)
        pickle.dump(stats, f)

        f.close()
